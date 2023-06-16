import csv
import json
import logging
import urllib.request
from io import BytesIO

from openpyxl import load_workbook

from main.models import Country, MetaData

logger = logging.getLogger(__name__)


# loads the resource from HDX. Only works with CSV currently
def _load_resource(url):
    try:
        return _load_csv(url)
    except UnicodeDecodeError:
        logger.info("Could not read csv, trying excel")
        return _load_excel(url)


def _load_csv(url):
    response = urllib.request.urlopen(url)
    lines = [line.decode("utf-8") for line in response.readlines()]
    csv_reader = csv.reader(lines)
    list_of_csv = list(csv_reader)
    logger.info("Successfully read csv")
    return list_of_csv


def _load_excel(url):
    return [
        list(row)
        for row in load_workbook(
            filename=BytesIO(urllib.request.urlopen(url).read()),
            data_only=True,
        ).active.iter_rows(values_only=True)
    ]


# takes in the header mapping, list of list data, and header rows and
# replaces headers
def _replace_headers(headerRow, mappings, data):
    headerRow = headerRow - 1
    for idx, header in enumerate(data[headerRow]):
        for mapping in mappings:
            if mapping[0] == header:
                data[headerRow][idx] = mapping[1]
    return data


# removes rows between header and data
def _remove_rows(headerRow, dataRow, data):
    return [data[headerRow - 1]] + data[dataRow - 1 :]


# add new columns as defined in fixed cols part of transformation file
def _add_fixed_values(fixedCols, data):
    for fixedCol in fixedCols:
        for idx, row in enumerate(data):
            if idx == 0:
                data[idx].append(fixedCol["key"])
            else:
                data[idx].append(fixedCol["value"])

    return data


def _add_metadata_and_country(data, iso3, hdx_id):
    country_obj = Country.objects.all().get(admin0_code_iso3=iso3)
    metadata_obj = MetaData.objects.all().get(hdx_id=hdx_id)
    data[0] += ["country", "metadata"]
    for row in data[1:]:
        row += [country_obj, metadata_obj]
    return data


def _transform_to_object(
    longDefinition, coreDefinition, countryDefinition, data
):
    longLookup = {}
    for definition in longDefinition:
        key = definition["header"]
        longLookup[key] = None
        for idx, header in enumerate(data[0]):
            if header == key:
                longLookup[key] = idx

    indicatorLookup = {}
    for definition in coreDefinition:
        key = definition["header"]
        indicatorLookup[key] = None
        for idx, header in enumerate(data[0]):
            if header == key:
                indicatorLookup[key] = idx

    for definition in countryDefinition:
        key = definition["header"]
        indicatorLookup[key] = None
        for idx, header in enumerate(data[0]):
            if header == key:
                indicatorLookup[key] = idx

    dbObj = []

    for idx, row in enumerate(data[1:]):
        for indicator in indicatorLookup:
            dbRow = {}
            if indicatorLookup[indicator] is not None:
                place = indicatorLookup[indicator]
                value = row[place]
                dbRow["key"] = indicator
                dbRow["value"] = value
            else:
                dbRow["key"] = indicator
                dbRow["value"] = None

            for key in longLookup:
                if longLookup[key] is not None:
                    place = longLookup[key]
                    value = row[place]
                    dbRow[key] = value
                else:
                    dbRow[key] = None

            dbRow["record_id"] = idx

            dbObj.append(dbRow)

    return dbObj


def transform(indicator):
    # schema for db
    longDefinitionFile = "definition_files/long_definition.json"

    with open(longDefinitionFile) as f:
        longDefinition = json.load(f)

    # schema for individual indicator segments
    indicatorDefinitionFile = (
        "definition_files/" + indicator + "_definition.json"
    )

    with open(indicatorDefinitionFile) as f:
        indicatorDefinition = json.load(f)

    dbObj = []

    # loop for each country in the indicator
    for country in indicatorDefinition["fields"]["country-specific"]:
        # load the transformation config to go from resource to DB schema
        for transformation in country["transformations"]:
            # get the resource and format
            data = _load_resource(transformation["resource"])

            # replace the headers to standardised headers
            data = _replace_headers(
                transformation["headerRow"], transformation["mappings"], data
            )

            # remove rows between and header and start of data
            data = _remove_rows(
                transformation["headerRow"], transformation["dataRow"], data
            )

            # add columns for fixed values
            data = _add_fixed_values(transformation["fixedCols"], data)
            # add foreign keys
            data = _add_metadata_and_country(
                data=data,
                iso3=country["country"],
                hdx_id=country["dataset_id"],
            )

            # transform to defined DB object from definition files
            newObj = _transform_to_object(
                longDefinition["fields"],
                indicatorDefinition["fields"]["core"],
                country["include"],
                data,
            )
            dbObj = dbObj + newObj

    print(dbObj[:10])

    return dbObj


if __name__ == "__main__":
    transform()
